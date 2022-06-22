from django.db.models import Sum
from django.db.models.query_utils import Q
from django.http.response import HttpResponse
from django.shortcuts import render
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from employee.models import AttendanceStatus, EmployeeAttendance, EmployeeLeave
from reports.forms import AttendanceReportForm, LeaverReportForm
from userRoles.views import getUserType
from django.contrib.auth.decorators import login_required, user_passes_test


def userHasAccess(user):

    if (getUserType(user) == "HR" or
            user.is_superuser):
        return True
    else:
        return False


@login_required
@user_passes_test(userHasAccess,
                  login_url="/dashboard/",
                  redirect_field_name=None)
def report_list(request):
    """Render a list with all reports available."""

    context = {
        "reports": {
            "Attendance Report": "reports:attendance_report",
            "Leaver Report": "reports:leaver_report"
        }
    }

    return render(request, 'reports/report_list.html', context)


@login_required
@user_passes_test(userHasAccess,
                  login_url="/dashboard/",
                  redirect_field_name=None)
def attendance_report(request):
    """Render a form with filter for the attendance report."""

    form = AttendanceReportForm()

    if request.method == "POST":
        form = AttendanceReportForm(request.POST)

    context = {
        "form": form
    }

    return render(request, 'reports/attendance_report.html', context)


@login_required
def leaver_report(request):
    """Render a form with filter for the leaver report."""

    form = LeaverReportForm()

    if request.method == "POST":
        form = LeaverReportForm(request.POST)

    context = {
        "form": form
    }

    return render(request, 'reports/leaver_report.html', context)


def get_employee_attendances(request):
    """Return a object with all employees and their attendances."""

    if request.method == "POST":

        filters = (Q(attendance_date__gte=request.POST.get("start_date")) &
                   Q(attendance_date__lte=request.POST.get("end_date")))

        employee_id = None if request.POST.get(
            'employee') == "" else request.POST.get('employee')
        department = None if request.POST.get(
            "department") == "" else request.POST.get("department")
        payroll_type = None if request.POST.get(
            "payroll_type") == "" else request.POST.get("payroll_type")

        if employee_id is not None:
            filters = filters & Q(employee_id=employee_id)
        if department is not None:
            filters = filters & Q(employee__department=department)
        if payroll_type is not None:
            filters = filters & Q(employee__payroll_type=payroll_type)

        employee_attendances = EmployeeAttendance.objects.filter(
            filters).prefetch_related('employee').order_by(
                "employee__first_name", "attendance_date").values(
                    "employee__first_name",
                    "employee__last_name",
                    "employee__id",
                    "employee__manager__first_name",
                    "employee__manager__last_name",
                    "attendance_date",
                    "status__status_name",
                    "hours_worked",
                    "additional_status__status_name",
                    "additional_hours_worked",
                    "notes"
        ).distinct()

        return employee_attendances


def get_employee_leave(request):
    """Return a object with all employees and their leave."""

    if request.method == "POST":

        filters = (Q(start_date__gte=request.POST.get("start_date")) &
                   Q(start_date__lte=request.POST.get("end_date")))

        employee_id = None if request.POST.get(
            'employee') == "" else request.POST.get('employee')
        department = None if request.POST.get(
            "department") == "" else request.POST.get("department")

        if employee_id is not None:
            filters = filters & Q(employee_id=employee_id)
        if department is not None:
            filters = filters & Q(employee__department=department)

        employees_leave = EmployeeLeave.objects.filter(
            filters).prefetch_related('employee').order_by(
                "start_date", "employee__first_name").values(
                    "employee__first_name",
                    "employee__last_name",
                    "employee__id",
                    "employee__manager__first_name",
                    "employee__manager__last_name",
                    "start_date",
                    "has_deduction",
                    "leave_type__status_name",
                    "notes"
        ).distinct()

        return employees_leave


def get_date_range(request):
    """Return a list with the date range beteween two dates."""

    start_date = datetime.strptime(
        request.POST.get("start_date"), "%Y-%m-%d")
    end_date = datetime.strptime(request.POST.get("end_date"), "%Y-%m-%d")

    date_list = []
    for d in range((end_date - start_date).days + 1):
        date_formated = datetime.strftime(
            start_date + timedelta(days=d), "%d/%m/%y")
        date_list.append(date_formated)

    return date_list


@login_required
def draw_leaver_report(request):
    """Draw spreadsheet with leavers."""

    today = datetime.today()
    employees_leave = get_employee_leave(request)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.cell(row=3, column=1).value = "First Name"
    worksheet.cell(row=3, column=2).value = "Last Name"
    worksheet.cell(row=3, column=3).value = "Manager"
    worksheet.cell(row=3, column=4).value = "Leave Date"
    worksheet.cell(row=3, column=5).value = "Deduction"
    worksheet.cell(row=3, column=6).value = "Leave Type"
    worksheet.cell(row=3, column=7).value = "Comments"

    for i, employee_leave in enumerate(employees_leave, start=4):
        worksheet.cell(
            row=i, column=1).value = employee_leave["employee__first_name"]
        worksheet.cell(
            row=i, column=2).value = employee_leave["employee__last_name"]

        if employee_leave["employee__manager__first_name"] is None:
            worksheet.cell(
                row=i, column=3).value = "-"
        else:
            worksheet.cell(
                row=i, column=3
            ).value = (employee_leave['employee__manager__first_name'] +
                       " " + employee_leave['employee__manager__last_name'])

        worksheet.cell(row=i, column=4).value = datetime.strftime(
            employee_leave["start_date"], "%d/%m/%Y")

        if employee_leave["has_deduction"]:
            worksheet.cell(row=i, column=5).value = "Yes"
            for j in range(worksheet.min_column, worksheet.max_column + 1):
                worksheet.cell(row=i, column=j).font = Font(
                    color="F94144", bold=True)
        else:
            worksheet.cell(row=i, column=5).value = "No"

        worksheet.cell(
            row=i, column=6).value = employee_leave["leave_type__status_name"]

        worksheet.cell(row=i, column=7).value = employee_leave["notes"]

    config_header_styles(worksheet, 3, worksheet.max_column)
    config_employee_styles(worksheet)

    worksheet.column_dimensions[get_column_letter(
        worksheet.max_column)].width = 150
    worksheet.column_dimensions[get_column_letter(6)].width = 40

    output = HttpResponse(save_virtual_workbook(
        workbook), content_type='application/ms-excel')
    output['Content-Disposition'] = f'attachment; \
        filename=leaver_report_{datetime.strftime(today, "%d/%m/%y")}.xlsx'

    return output


@login_required
@user_passes_test(userHasAccess,
                  login_url="/dashboard/",
                  redirect_field_name=None)
def draw_attendance_report(request):
    """Draw a spreadsheet with the attendance report."""

    workbook = Workbook()
    worksheet = workbook.active

    today = datetime.today()
    employee_attendances = get_employee_attendances(request)
    employees = employee_attendances.order_by(
        "employee__first_name").values_list(
        "employee_id",
        "employee__first_name",
        "employee__last_name",
        "employee__manager__first_name",
        "employee__manager__last_name",
        "employee__employee_status"
    ).distinct()
    date_range = get_date_range(request)
    status = AttendanceStatus.objects.all()
    step = status.count()

    # Freeze panes for the headers/rows
    worksheet.freeze_panes = "C4"

    # Set the first col and row.
    last_col_num = 9
    first_row = 4

    # Set Headers
    worksheet.cell(row=3, column=1).value = "First Name"
    worksheet.cell(row=3, column=2).value = "Last Name"
    worksheet.cell(row=3, column=3).value = "Manager"
    worksheet.cell(row=3, column=4).value = "Active Total"
    worksheet.cell(row=3, column=5).value = "Inactive Total"
    worksheet.cell(row=3, column=6).value = "Holidays Total"
    worksheet.cell(row=3, column=7).value = "Bank Holidays Total"
    worksheet.cell(row=3, column=8).value = "Clawback Total"
    worksheet.cell(row=3, column=9).value = "Comments"

    config_header_styles(worksheet, 3, last_col_num)

    worksheet.column_dimensions[get_column_letter(9)].width = 50

    for employee in employees:
        employee_first_name = employee[1]
        employee_last_name = employee[2]
        employee_status = employee[5]
        manager_name = ""

        # Config and add employees name
        if employee_status == "Leaver":
            worksheet.cell(row=first_row, column=1).font = Font(color="F94144")
            worksheet.cell(row=first_row, column=2).font = Font(color="F94144")
        worksheet.cell(
            row=first_row, column=1).value = employee_first_name
        worksheet.cell(
            row=first_row, column=2).value = employee_last_name

        # Config and add managers name
        if (employee[3] or employee[4]) is None:
            manager_name = "-"
        else:
            manager_name = employee[3] + " " + employee[4]

        worksheet.cell(
            row=first_row, column=3).value = manager_name

        # Config and add total active hours
        worksheet.cell(
            row=first_row, column=4).value = get_active_hours(
                employee[0], employee_attendances)

        # Config and add total inactive hours
        worksheet.cell(
            row=first_row, column=5).value = get_inactive_hours(
                employee[0], employee_attendances)

        # Config and add total holidays hours
        worksheet.cell(
            row=first_row, column=6).value = get_holidays_hours(
                employee[0], employee_attendances)

        # Config and add total bank holidays hours
        worksheet.cell(
            row=first_row, column=7).value = get_bank_holidays_hours(
                employee[0], employee_attendances)

        # Config and add total clawback hours
        worksheet.cell(
            row=first_row, column=8).value = get_clawback_hours(
                employee[0], employee_attendances)

        # Add comments
        worksheet.cell(
            row=first_row, column=9).value = get_attendance_comments(
                employee[0],
            employee_attendances)

        first_row += 1

    # Config and add the data
    write_employee_attendance_data(worksheet, employee_attendances,
                                   date_range, last_col_num,
                                   step, employees, status)

    # Config styles for status
    config_status_styles(worksheet)

    # config styles for employees
    config_employee_styles(worksheet)

    output = HttpResponse(save_virtual_workbook(
        workbook), content_type='application/ms-excel')
    output['Content-Disposition'] = f'attachment; \
        filename=attendance_report_{datetime.strftime(today, "%d/%m/%y")}.xlsx'

    return output


def get_active_hours(employee_id, employee_attendances):
    """ Returns all active hours for an employee """

    # Query the total active hours.
    normal_active_hours = employee_attendances.filter(
        Q(employee_id=employee_id) &
        (Q(status__status_name="Active") |
         Q(status__status_name="Inactive paid") |
         Q(status__status_name="Induction"))
    ).values_list(
        "attendance_date",
        flat=True,
    ).distinct().aggregate(
        normal_active_hours=Sum("hours_worked"))["normal_active_hours"]

    # Query the total additional active hours.
    ad_active_hours = employee_attendances.filter(
        Q(employee_id=employee_id) &
        (Q(additional_status__status_name="Active") |
         Q(additional_status__status_name="Inactive paid") |
         Q(additional_status__status_name="Induction"))
    ).values_list(
        "attendance_date",
        flat=True,
    ).distinct().aggregate(
        ad_active_hours=Sum("additional_hours_worked"))["ad_active_hours"]

    total_active_hours = (0 if normal_active_hours is None
                          else normal_active_hours) + \
        (0 if ad_active_hours is None else ad_active_hours)

    return total_active_hours


def get_inactive_hours(employee_id, employee_attendances):
    """ Returns all inactive hours for an employee """

    # Query the total inactive hours.
    normal_inactive_hours = employee_attendances.filter(
        Q(employee_id=employee_id) &
        (Q(status__status_name="Inactive unpaid") |
         Q(status__status_name="Sick") |
         Q(status__status_name="AWOL") |
         Q(status__status_name="Maternity Leave"))
    ).values_list(
        "attendance_date",
        flat=True,
    ).distinct().aggregate(
        normal_inactive_hours=Sum("hours_worked"))["normal_inactive_hours"]

    # Query the total additional inactive hours.
    ad_inactive_hours = employee_attendances.filter(
        Q(employee_id=employee_id) &
        (Q(additional_status__status_name="Inactive unpaid") |
         Q(additional_status__status_name="Sick") |
         Q(additional_status__status_name="AWOL") |
         Q(additional_status__status_name="Maternity Leave"))
    ).values_list(
        "attendance_date",
        flat=True,
    ).distinct().aggregate(
        ad_inactive_hours=Sum("additional_hours_worked"))["ad_inactive_hours"]

    total_inactive_hours = (0 if normal_inactive_hours is None
                            else normal_inactive_hours) + \
        (0 if ad_inactive_hours is None else ad_inactive_hours)

    return total_inactive_hours


def get_holidays_hours(employee_id, employee_attendances):
    """ Returns all holiday hours for an employee """

    # Query the total holiday hours.
    holidays_hours = employee_attendances.filter(
        Q(employee_id=employee_id) &
        Q(status__status_name="Holidays")
    ).values_list(
        "attendance_date",
        flat=True,
    ).distinct().aggregate(
        holidays_hours=Sum("hours_worked"))["holidays_hours"]

    ad_holidays_hours = employee_attendances.filter(
        Q(employee_id=employee_id) &
        (Q(additional_status__status_name="Holidays"))
    ).values_list(
        "attendance_date",
        flat=True,
    ).distinct().aggregate(
        ad_holidays_hours=Sum("additional_hours_worked"))["ad_holidays_hours"]

    total_holidays_hours = (0 if holidays_hours is None
                            else holidays_hours) + \
        (0 if ad_holidays_hours is None else ad_holidays_hours)

    return total_holidays_hours


def get_bank_holidays_hours(employee_id, employee_attendances):
    """ Returns all bank holiday hours for an employee """

    # Query the total bank holiday hours.
    bank_holidays_hours = employee_attendances.filter(
        Q(employee_id=employee_id) &
        Q(status__status_name="Bank Holiday")
    ).values_list(
        "attendance_date",
        flat=True,
    ).distinct().aggregate(
        bank_holidays_hours=Sum("hours_worked"))["bank_holidays_hours"]

    ad_bank_holidays_hours = employee_attendances.filter(
        Q(employee_id=employee_id) &
        (Q(additional_status__status_name="Bank Holiday"))
    ).values_list(
        "attendance_date",
        flat=True,
    ).distinct().aggregate(
        ad_bank_holidays_hours=Sum("additional_hours_worked")
    )["ad_bank_holidays_hours"]

    total_bank_holidays_hours = (0 if bank_holidays_hours is None
                                 else bank_holidays_hours) + \
        (0 if ad_bank_holidays_hours is None else ad_bank_holidays_hours)

    return total_bank_holidays_hours


def get_clawback_hours(employee_id, employee_attendances):
    """ Returns all clawback hours for an employee """

    # Query the total clawback hours.
    clawback_hours = employee_attendances.filter(
        Q(employee_id=employee_id) &
        Q(status__status_name="Clawback")
    ).values_list(
        "attendance_date",
        flat=True,
    ).distinct().aggregate(
        clawback_hours=Sum("hours_worked"))["clawback_hours"]

    ad_clawback_hours = employee_attendances.filter(
        Q(employee_id=employee_id) &
        (Q(additional_status__status_name="Clawback"))
    ).values_list(
        "attendance_date",
        flat=True,
    ).distinct().aggregate(
        ad_clawback_hours=Sum("additional_hours_worked"))["ad_clawback_hours"]

    total_clawback_hours = (0 if clawback_hours is None
                            else clawback_hours) + \
        (0 if ad_clawback_hours is None else ad_clawback_hours)

    return total_clawback_hours


def get_attendance_comments(employee_id, employee_attendances):
    """ Returns a concatenation of all comments with date. """

    notes = ""

    attendance_comments = employee_attendances.filter(
        employee_id=employee_id).values_list("attendance_date", "notes")

    for comment in attendance_comments:
        if comment[1] != "":
            notes += comment[0].strftime("%d/%m/%Y") + \
                ": " + comment[1] + " | "

    return notes


def config_employee_styles(worksheet):
    """ Configures the employee styles """

    for row in worksheet.iter_rows(min_row=4, max_row=worksheet.max_row,
                                   min_col=1, max_col=worksheet.max_column):
        for cell in row:
            worksheet[(cell.coordinate)].border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin'))
            worksheet[cell.coordinate].alignment = Alignment(
                horizontal="center")
            if cell.row % 2 == 0:
                cell.fill = PatternFill(
                    fill_type="solid",
                    start_color="D3D3D3")


def config_status_styles(worksheet):
    """ Configures the status styles """

    for row in worksheet.iter_rows(min_row=2, max_row=2,
                                   min_col=10, max_col=worksheet.max_column):
        for cell in row:
            cell.font = Font(bold=True)
            worksheet.column_dimensions[get_column_letter(
                cell.column)].width = 5
            cell.alignment = Alignment(
                text_rotation=90,
                vertical="center",
                horizontal="center")
            cell.border = Border(
                left=Side(style='medium'),
                right=Side(style='medium'),
                top=Side(style='medium'),
                bottom=Side(style='medium'))


def config_header_styles(worksheet, start_col, end_col):
    """ Configures the header styles """

    for row in worksheet.iter_rows(min_row=start_col, max_row=start_col,
                                   min_col=1, max_col=end_col):
        for cell in row:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            worksheet.column_dimensions[get_column_letter(
                cell.column)].width = 20
            cell.border = Border(
                left=Side(style='medium'),
                right=Side(style='medium'),
                top=Side(style='medium'),
                bottom=Side(style='medium'))


def write_employee_attendance_data(worksheet, employee_attendances,
                                   date_range, last_col_num,
                                   step, employees, status):
    """ Writes the employee attendance data """

    for date in date_range:
        # Get columns letter
        next_first_col = get_column_letter(last_col_num + 1)
        last_col = get_column_letter(last_col_num + step)

        # Merge the date cells and writes the date
        worksheet.merge_cells(f"{next_first_col}1:{last_col}1")
        worksheet[f"{next_first_col}1"] = date
        worksheet[f"{next_first_col}1"].alignment = Alignment(
            horizontal="center")
        worksheet[f"{next_first_col}1"].border = Border(
            left=Side(style='medium'),
            right=Side(style='medium'),
            top=Side(style='medium'),
            bottom=Side(style='medium'))
        worksheet[f"{next_first_col}1"].font = Font(bold=True, size=16)
        worksheet[f"{next_first_col}1"].fill = PatternFill(
            fill_type="solid",
            start_color="43AA8B")

        # Writes the statuses
        for y, s in enumerate(status, start=last_col_num + 1):
            worksheet[f"{get_column_letter(y)}2"] = s.status_name

            # Writes the hours
            for x, e in enumerate(employees, start=4):
                for emp in employee_attendances:
                    if (emp["employee__id"] == e[0]
                        and emp["status__status_name"] == s.status_name
                            and datetime.strftime(
                                emp["attendance_date"], "%d/%m/%y") == date):
                        worksheet[
                            f"{get_column_letter(y)}{x}"] = emp["hours_worked"]
                    elif (emp["employee__id"] == e[0]
                          and emp["additional_status__status_name"] == s.status_name
                            and datetime.strftime(
                                emp["attendance_date"], "%d/%m/%y") == date):
                        worksheet[
                            f"{get_column_letter(y)}{x}"
                        ] = emp["additional_hours_worked"]
        last_col_num = last_col_num + step
