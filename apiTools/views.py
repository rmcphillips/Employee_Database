from django.http.response import HttpResponse
from django.shortcuts import render
from .forms import DncApiForm
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from datetime import datetime
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.styles import Font
import json
import xmltodict
import csv
import os
import requests

# ! Initialize the list to be used on the exports
responseMessage = []

# ! Gets the current system date to be used on the exports
today = datetime.now()


def userHasAccess(user):
    if user.department.department_name == "Technology":
        return True
    else:
        return False


@login_required
@user_passes_test(userHasAccess,
                  login_url="/dashboard/",
                  redirect_field_name=None)
def dncApi(request):
    """ Delete customer from campaign. """

    form = DncApiForm()

    responseMessage.clear()

    if request.method == "POST":

        form = DncApiForm(request.POST)

        if form.is_valid():

            campaignID = request.POST["campaignID"]
            customerList = list(request.POST["customerContactNumber"].split())
            tenant = request.POST["tenant"]
            apiKey = getApiKey(tenant)
            server = getServer(tenant)
            customerIDList = getCustomerByPhone(
                customerList, campaignID, tenant)

            for currentCustomer in customerIDList:

                url = f"{server}/api/tstats/campaigns/{campaignID}/customers/{currentCustomer}"

                payload = {}
                headers = {
                    'Authorization': apiKey
                }

                response = requests.request(
                    "DELETE", url, headers=headers, data=payload)

                if response.status_code == 401:
                    messages.warning(request, f"""
                    Status {response.status_code}
                        - Authentication Error.""")
                    break

                elif (response.status_code == 200 or
                      response.status_code == 204):

                    responseMessage.append(
                        {
                            "tenant": tenant,
                            "campaignID": campaignID,
                            "customerID": currentCustomer,
                            "status": response.status_code,
                            "message": f"Customer ID: {currentCustomer} deleted successfully.",
                            "date": today.strftime("%c")
                        })
                else:
                    messageParsed = json.loads(response.content)
                    messageParsed = messageParsed["message"]

                    responseMessage.append(
                        {
                            "tenant": tenant,
                            "campaignID": campaignID,
                            "customerID": currentCustomer,
                            "status": response.status_code,
                            "message": f"Customer ID: {currentCustomer} - {messageParsed}",
                            "date": today.strftime("%c")
                        })

    context = {
        "form": form,
        "responseMessage": responseMessage
    }

    return render(request, "apiTools/dncApi.html", context)


def getApiKey(tenant):
    """
        Return API key based on tenant.

        Required Arguments:
        tenant
    """

    if tenant == "iber":

        authKey = os.environ.get("IBERDROLA_API_KEY")

    elif tenant == "inm":

        authKey = os.environ.get("INM_API_KEY")

    elif tenant == "pin":

        authKey = os.environ.get("PINERGY_API_KEY")

    return authKey


def dncApiExportCSV(request):
    """ Exports current query to CSV file """
    campaignID = responseMessage[0]["campaignID"]
    response = HttpResponse(content_type="text/csv")
    response['Content-Disposition'] = f'attachment; \
        filename=dncReport_{campaignID}_{today.strftime("%x")}".csv"'

    if responseMessage:

        writer = csv.writer(response)
        writer.writerow(
            ["Tenant", "Campaign ID",
             "Customer ID", "Status", "Message", "Date"])

        for r in responseMessage:
            if r["tenant"]:
                r["tenant"] = getTenantName(r["tenant"])
            writer.writerow(
                r.values())

    return response


def dncApiExportXLS(request):
    """ Exports current query to XLS file """

    excelData = responseMessage
    wb = Workbook()
    ws = wb.active

    ws["A1"].value = "Tenant"
    ws["B1"].value = "Campaign ID"
    ws["C1"].value = "Customer ID"
    ws["D1"].value = "Status"
    ws["E1"].value = "Message"
    ws["F1"].value = "Date"

    for cell in ws["1:1"]:
        cell.font = Font(bold=True)

    row = 2
    for r in excelData:
        ws[f"A{row}"].value = getTenantName(r["tenant"])
        ws[f"B{row}"].value = r["campaignID"]
        ws[f"C{row}"].value = r["customerID"]
        ws[f"D{row}"].value = r["status"]
        ws[f"E{row}"].value = r["message"]
        ws[f"F{row}"].value = r["date"]
        row += 1

        response = HttpResponse(save_virtual_workbook(
            wb), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = f'attachment; \
            filename=dncReport_{r["campaignID"]}_{today.strftime("%x")}".xlsx"'

    return response


def getTenantName(tenant):
    """
    Returns correct tenant name.
    Required arguments: tenant
    """

    if tenant == "iber":
        tenant = "Iberdrola"
    elif tenant == "inm":
        tenant = "Independent News & Media"
    elif tenant == "pin":
        tenant = "Pinergy"

    return tenant


def getServer(tenant):
    """
    Returns correct server based on tenant passed.
    """

    if tenant == "iber":
        return "https://vcc-eu8.8x8.com"
    else:
        return "https://vcc-eu9.8x8.com"


def getCustomerByPhone(phoneList, campaignID, tenant):
    """
    Return customerID by phoneNumber provided.
    It receives a list of phone numbers.
    Required Argument: phoneList
    """

    customerList = []
    server = getServer(tenant)

    for phoneNumber in phoneList:
        url = f"{server}/api/stats/campaigns/{campaignID}/records?phone-list=*Contact%20Number%7C{phoneNumber}"

        payload = {}
        headers = {
            'Authorization': "Basic c2FsZXNzZW5zZWliMDE6ZDdiYmRmZGYyM2ZlY2M1ZDBiNDRlNjA1YmE2ZGZkNmMyMGZmZmRmMmYxZmUwMTQ1ZWU3ZmU4MzkzZGYxNjg1Ng=="
        }

        response = requests.request(
            "GET", url, headers=headers, data=payload)

        obj = xmltodict.parse(response.content)

        if not obj["records"] is None:
            customerList.append(obj["records"]["record"]["record-id"])
        else:
            customerList.append(phoneNumber)

    return customerList
